"""Extracteurs XLSX (openpyxl) et CSV (stdlib + chardet).

Format de sortie : une section `## Feuille : Nom` par feuille (ou par fichier pour CSV),
avec une table Markdown standard (entête + séparateur `---` + lignes).
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dépendance optionnelle absente
    load_workbook = None  # type: ignore[assignment]

from trimtokens.exceptions import MissingDependencyError
from trimtokens.extractors._encoding import detect_encoding
from trimtokens.models import ExtractedDocument, ExtractOptions, Section


def extract(path: Path, options: ExtractOptions) -> ExtractedDocument:
    ext = path.suffix.lower()
    if ext == ".csv":
        # CSV utilise stdlib `csv` — pas de dépendance externe requise.
        return _extract_csv(path)
    return _extract_xlsx(path)


def _extract_xlsx(path: Path) -> ExtractedDocument:
    if load_workbook is None:
        raise MissingDependencyError(
            "Le package 'openpyxl' est requis pour extraire les fichiers .xlsx. "
            "Installez l'extra : pip install 'trimtokens[office]'"
        )
    workbook = load_workbook(path, read_only=True, data_only=True)
    sections: list[Section] = []
    try:
        for sheet in workbook.worksheets:
            rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
            if not rows or all(_is_empty_row(r) for r in rows):
                continue
            markdown_table = _rows_to_markdown_table(rows)
            sections.append(Section(header=f"Feuille : {sheet.title}", content=markdown_table))
    finally:
        workbook.close()

    return ExtractedDocument(
        source_path=path,
        source_type="xlsx",
        sections=sections,
    )


def _extract_csv(path: Path) -> ExtractedDocument:
    raw = path.read_bytes()
    encoding = detect_encoding(raw)
    text = raw.decode(encoding, errors="replace")

    rows: list[tuple[Any, ...]] = [tuple(row) for row in csv.reader(text.splitlines())]
    if not rows:
        return ExtractedDocument(
            source_path=path,
            source_type="csv",
            sections=[],
            metadata={"encoding": encoding},
        )

    markdown_table = _rows_to_markdown_table(rows)
    return ExtractedDocument(
        source_path=path,
        source_type="csv",
        sections=[Section(header=f"Feuille : {path.stem}", content=markdown_table)],
        metadata={"encoding": encoding},
    )


def _rows_to_markdown_table(rows: list[tuple[Any, ...]]) -> str:
    if not rows:
        return ""

    width = max(len(row) for row in rows)
    header_row = list(rows[0]) + [""] * (width - len(rows[0]))
    body_rows = rows[1:]

    lines: list[str] = []
    lines.append("| " + " | ".join(_cell_to_str(c) for c in header_row) + " |")
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for row in body_rows:
        padded = list(row) + [""] * (width - len(row))
        lines.append("| " + " | ".join(_cell_to_str(c) for c in padded) + " |")
    return "\n".join(lines)


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)
